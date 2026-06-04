import io
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

def get_headers_and_rows(data, section):
    if section == "all":
        rows_data = data.get("recent_highlights", [])
        headers = ["Client", "Project", "Revenue", "Expenses", "Profit", "Margin %", "Status", "Outstanding"]
        rows = []
        for r in rows_data:
            rows.append([
                str(r.get("client_name", "")),
                str(r.get("project_name", "")),
                str(r.get("total_revenue", "")),
                str(r.get("total_expenses", "")),
                str(r.get("profit", "")),
                str(r.get("profit_margin_percent", "")),
                str(r.get("latest_invoice_status", "")),
                str(r.get("outstanding_amount", ""))
            ])
        return headers, rows

    rows_data = data.get("rows", [])
    if section == "financial":
        headers = ["Invoice No", "Date", "Client", "Total", "Paid", "Balance", "Status", "Due Date", "Days Overdue"]
        rows = []
        for r in rows_data:
            rows.append([
                str(r.get("invoice_no", "")),
                str(r.get("invoice_date", "")),
                str(r.get("client_name", "")),
                str(r.get("invoice_total", "")),
                str(r.get("amount_paid", "")),
                str(r.get("outstanding_balance", "")),
                str(r.get("status", "")),
                str(r.get("due_date", "")),
                str(r.get("days_overdue", ""))
            ])
        return headers, rows

    if section == "project":
        headers = ["Project No", "Project", "Status", "Budget", "Invoiced", "Received", "Expenses", "Profit"]
        rows = []
        for r in rows_data:
            rows.append([
                str(r.get("project_no", "")),
                str(r.get("project_name", "")),
                str(r.get("project_status", "")),
                str(r.get("budget", "")),
                str(r.get("invoiced", "")),
                str(r.get("received", "")),
                str(r.get("expenses", "")),
                str(r.get("profit", ""))
            ])
        return headers, rows

    if section == "payment":
        headers = ["Payment Date", "Invoice No", "Client", "Method", "Amount", "Invoice Status"]
        rows = []
        for r in rows_data:
            rows.append([
                str(r.get("payment_date", "")),
                str(r.get("invoice__invoice_no", "")),
                str(r.get("invoice__client__company_name", "")),
                str(r.get("payment_method", "")),
                str(r.get("amount", "")),
                str(r.get("invoice__status", ""))
            ])
        return headers, rows

    if section == "po-invoice":
        headers = ["PO No", "Vendor", "Total Amount", "Paid", "Balance"]
        rows = []
        for r in rows_data:
            rows.append([
                str(r.get("po_no", "")),
                str(r.get("vendor__name", "")),
                str(r.get("total_amount", "")),
                str(r.get("paid", "")),
                str(r.get("balance", ""))
            ])
        return headers, rows

    return [], []

def generate_excel_from_data(data, section):
    headers, rows = get_headers_and_rows(data, section)
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin

    # Rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_from_data(data, section):
    headers, rows = get_headers_and_rows(data, section)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    
    table_data = [headers] + rows
    if not table_data or len(table_data) == 1:
        table_data.append(["No data available"] * len(headers))

    t = Table(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    doc.build([t])
    output.seek(0)
    return output
